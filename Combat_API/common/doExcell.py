# -*-coding:utf-8-*-
# @Time      :2019/3/10/024 16:36
# @Author    :Tanxi
# @Email     :1410510771@qq.com
# @File      :openpyxl_0310.py
# @Software  :PyCharm Community Edition

# 1、从Excel读取数据，并存入列表；
# 2、写回Excel中实际结果，以及是否通过；
# 3、对Excel中Params列字段中的'tel'进行替换，替换tel表单中的电话号码，若已经存在电话则更新tel+1；
# 4、对从配置文件中读取到的需运行的case_id，进行判断

from openpyxl import load_workbook

from Combat_API.common import project_path
from Combat_API.common.python_logging import MyLogger
from Combat_API.common.read_config import ReadConfig

log = MyLogger()


class PyExcel:
    '''类的作用：是完成Excel数据的读，写，新建表单的操作'''

    def __init__(self, file, sheetname):
        self.file = file
        self.sheetname = sheetname

    def read_data(self, option):
        '''
        将Excel中指定表单中每行数据存为一个列表row_list,再将其作为子列表存入all_row_list中
        :param file: 文件名
        :param sheet: 表单名
        :return: all_row_list--一个嵌套列表
        '''

        case_id = ReadConfig(project_path.conf_path).read_str(option, 'case_id')

        log.info('开始读取数据:')
        all_row_list = []  # 把每行数据作为子列表，存放在该列表中
        try:
            wb = load_workbook(self.file)  # 打开excel文件
            sheet = wb[self.sheetname]  # 定位表单
        except Exception as e:
            log.error('读取文件错误，请检查文件名或表单名是否正确!报错信息:{}'.format(e))
            raise e

        tel = self.read_tel()  # 获取excel里面的电话号码

        # 对应测试用例中用例编号
        if case_id == 'all':
            row_case = range(1, sheet.max_row)
        else:
            row_case = eval(case_id)
        for item in row_case:
            row_list = []  # 存放每行单元格内数据，每一行先将其置空
            for i in (1, 2, 3, 5, 6, 7, 8):  # id，模块名，接口名，请求方法，参数，(sql)，期望结果。[没有sql的表单，第7列对应期望结果]
                row_list.append(sheet.cell(row=item + 1, column=i).value)  # 用例id+1才是需读取数据的对应行数
            if row_list[4].find('tel') != -1:  # 列表中参数列，当参数中可以找到tel字段
                row_list[4] = row_list[4].replace('tel', str(tel))  # 替换值 tel  replace只能用于str替换
            else:
                self.update_tel(tel + 1)
            all_row_list.append(row_list)
            # 先全部读取数据，再返回需执行的对应case_id数据
            # final_data = []
            # if case_id == 'all':
            #     final_data = all_row_list
            # else:
            #     for j in case_id:
            #         final_data.append(all_row_list[j-1])

        log.info('读取到的数据为：{}'.format(all_row_list))
        log.info('读取数据完毕')
        wb.close()
        return all_row_list

    def read_tel(self):
        '''获取存在Excel里面的tel'''
        wb = load_workbook(self.file)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1, 2).value  # 返回电话号码的值

    def write_data(self, row, column, value):
        '''
        往指定Excel文件的指定表单指定行列插入指定数据
        :param file: 文件名
        :param sheet: 表单名
        :param row: 表单行
        :param column: 表单列
        :param value: 要写入的数据
        :return: None
        '''
        log.info('开始写入数据:')
        try:
            wb = load_workbook(self.file)  # 打开excel文件
            sheet = wb[self.sheetname]  # 定位表单
        except Exception as e:
            log.error('写入数据出错，请检查文件名及相关信息是否正确。错误信息：{}'.format(e))
            raise e
        sheet.cell(row, column).value = value  # 将值写入指定单元格
        wb.save(self.file)

        log.info('写入的数据为：{}'.format(value))
        log.info('写入数据完毕')
        wb.close()  # 写文件后一定记得关闭

    def update_tel(self, new_tel):
        '''写回手机号码'''
        wb = load_workbook(self.file)
        sheet = wb['tel']
        sheet.cell(1, 2, new_tel)
        wb.save(self.file)
        wb.close()


if __name__ == '__main__':
    t = PyExcel(project_path.case_path, 'recharge')
    # a = t.read_tel()
    #     print(t.read_data())
    #     t.write_data(38,12,'test')
    a = t.read_data('rechargeCase')
    print(a)
    b=a[1][6]
    # c=b['LeaveAmount']
    # print(c)
    print(type(b))
